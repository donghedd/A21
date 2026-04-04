"""
Excel File Loader
"""
from typing import List
from .base import BaseLoader, Document


class ExcelLoader(BaseLoader):
    """Load Excel files using openpyxl"""
    
    def load(self, file_path: str) -> List[Document]:
        """Load Excel file and extract text from all sheets"""
        try:
            from openpyxl import load_workbook
            
            documents = []
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []
                
                for row in sheet.iter_rows(values_only=True):
                    # Filter out empty cells
                    row_values = [str(cell) if cell is not None else '' for cell in row]
                    if any(v.strip() for v in row_values):
                        rows.append(' | '.join(row_values))
                
                if rows:
                    content = '\n'.join(rows)
                    documents.append(Document(
                        page_content=content,
                        metadata={
                            'source': file_path,
                            'sheet': sheet_name,
                            'file_type': 'xlsx'
                        }
                    ))
            
            wb.close()
            return documents
            
        except Exception as e:
            raise Exception(f"Failed to load Excel file: {str(e)}")
